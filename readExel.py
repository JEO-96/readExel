from openpyxl import load_workbook
from openpyxl import Workbook

load_wb = load_workbook("./text.xlsx", data_only=True)

load_ws = load_wb['미딕스-팝송']

write_wb = Workbook()
write_ws = write_wb.active
str1 = ''
for i in range(1, 1013):
    write_ws.append([0])
for i in range(1013, 1274):
    str1 = load_ws.cell(i, 2).value
    if '-' in str1:
        str1 = str1.split('-')
        print(str1[0] + '!' + str1[1])
        write_ws.append([str1[0], str1[1]])
    else:
        print(str1)
        write_ws.append([str1])

write_wb.save('./output.xlsx')
